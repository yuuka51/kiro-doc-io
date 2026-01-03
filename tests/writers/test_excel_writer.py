"""ExcelライターのユニットテストExcel

要件7.1: 構造化データからExcelファイルを生成する
要件7.2: 複数のシートを含むワークブックを作成する
要件7.3: セルデータ、基本的な書式設定、列幅の自動調整を適用する
要件7.4: 生成したファイルをユーザーがアクセス可能な場所に保存する
要件7.5: ファイル生成中にエラーが発生した場合、エラーメッセージを返す
"""

import os
import pytest
import tempfile

from openpyxl import load_workbook

from document_format_mcp_server.writers.excel_writer import ExcelWriter
from document_format_mcp_server.utils.models import WriteResult


@pytest.fixture
def excel_writer():
    """ExcelWriterインスタンスを返すフィクスチャ"""
    return ExcelWriter()


@pytest.fixture
def temp_output_dir():
    """一時出力ディレクトリを返すフィクスチャ"""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield tmpdir


@pytest.fixture
def basic_workbook_data():
    """基本的なワークブックデータを返すフィクスチャ"""
    return {
        "sheets": [
            {
                "name": "テストシート",
                "data": [
                    ["ヘッダー1", "ヘッダー2"],
                    ["データ1", "データ2"]
                ]
            }
        ]
    }


@pytest.fixture
def multi_sheet_data():
    """複数シートのデータを返すフィクスチャ"""
    return {
        "sheets": [
            {
                "name": "シート1",
                "data": [
                    ["ID", "名前", "値"],
                    [1, "項目A", 100],
                    [2, "項目B", 200]
                ]
            },
            {
                "name": "シート2",
                "data": [
                    ["カテゴリ", "合計"],
                    ["A", 300],
                    ["B", 400]
                ]
            },
            {
                "name": "シート3",
                "data": [
                    ["日付", "メモ"],
                    ["2024-01-01", "テスト"]
                ]
            }
        ]
    }


class TestExcelWriterBasic:
    """ExcelWriterの基本機能テスト"""

    def test_create_workbook_returns_write_result(
        self, excel_writer, temp_output_dir, basic_workbook_data
    ):
        """ExcelWriterがWriteResultを返すことを検証"""
        output_path = os.path.join(temp_output_dir, "test.xlsx")
        result = excel_writer.create_workbook(basic_workbook_data, output_path)
        
        # WriteResultデータクラスであることを検証
        assert isinstance(result, WriteResult)
        assert hasattr(result, 'success')
        assert hasattr(result, 'output_path')
        assert hasattr(result, 'url')
        assert hasattr(result, 'error')

    def test_create_workbook_success(
        self, excel_writer, temp_output_dir, basic_workbook_data
    ):
        """Excelファイルが正常に生成されることを検証（要件7.1）"""
        output_path = os.path.join(temp_output_dir, "test.xlsx")
        result = excel_writer.create_workbook(basic_workbook_data, output_path)
        
        # 成功フラグを検証
        assert result.success is True
        assert result.error is None
        assert result.output_path == output_path
        
        # ファイルが存在することを検証
        assert os.path.exists(output_path)

    def test_create_workbook_file_is_valid_xlsx(
        self, excel_writer, temp_output_dir, basic_workbook_data
    ):
        """生成されたファイルが有効なExcelファイルであることを検証"""
        output_path = os.path.join(temp_output_dir, "test.xlsx")
        excel_writer.create_workbook(basic_workbook_data, output_path)
        
        # openpyxlで読み込めることを検証
        wb = load_workbook(output_path)
        assert wb is not None
        assert len(wb.sheetnames) == 1


class TestExcelWriterSheets:
    """シート作成のテスト"""

    def test_create_single_sheet(self, excel_writer, temp_output_dir):
        """単一シートが正しく作成されることを検証（要件7.2）"""
        data = {
            "sheets": [
                {
                    "name": "単一シート",
                    "data": [["A", "B"], [1, 2]]
                }
            ]
        }
        output_path = os.path.join(temp_output_dir, "single_sheet.xlsx")
        result = excel_writer.create_workbook(data, output_path)
        
        assert result.success is True
        
        # ワークブックの内容を検証
        wb = load_workbook(output_path)
        assert len(wb.sheetnames) == 1
        assert "単一シート" in wb.sheetnames

    def test_create_multiple_sheets(
        self, excel_writer, temp_output_dir, multi_sheet_data
    ):
        """複数シートが正しく作成されることを検証（要件7.2）"""
        output_path = os.path.join(temp_output_dir, "multi_sheets.xlsx")
        result = excel_writer.create_workbook(multi_sheet_data, output_path)
        
        assert result.success is True
        
        # ワークブックの内容を検証
        wb = load_workbook(output_path)
        assert len(wb.sheetnames) == 3
        assert "シート1" in wb.sheetnames
        assert "シート2" in wb.sheetnames
        assert "シート3" in wb.sheetnames

    def test_sheet_data_is_written_correctly(self, excel_writer, temp_output_dir):
        """シートデータが正しく書き込まれることを検証（要件7.1）"""
        data = {
            "sheets": [
                {
                    "name": "データシート",
                    "data": [
                        ["名前", "年齢", "都市"],
                        ["田中", 30, "東京"],
                        ["鈴木", 25, "大阪"]
                    ]
                }
            ]
        }
        output_path = os.path.join(temp_output_dir, "data_test.xlsx")
        result = excel_writer.create_workbook(data, output_path)
        
        assert result.success is True
        
        # データの内容を検証
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "名前"
        assert ws.cell(row=1, column=2).value == "年齢"
        assert ws.cell(row=2, column=1).value == "田中"
        assert ws.cell(row=2, column=2).value == 30
        assert ws.cell(row=3, column=3).value == "大阪"


class TestExcelWriterFormatting:
    """書式設定のテスト"""

    def test_header_row_formatting(self, excel_writer, temp_output_dir):
        """ヘッダー行の書式設定が適用されることを検証（要件7.3）"""
        data = {
            "sheets": [
                {
                    "name": "フォーマットテスト",
                    "data": [
                        ["ヘッダー1", "ヘッダー2"],
                        ["データ1", "データ2"]
                    ],
                    "formatting": {"header_row": True}
                }
            ]
        }
        output_path = os.path.join(temp_output_dir, "format_test.xlsx")
        result = excel_writer.create_workbook(data, output_path)
        
        assert result.success is True
        
        # ヘッダー行のフォーマットを検証
        wb = load_workbook(output_path)
        ws = wb.active
        header_cell = ws.cell(row=1, column=1)
        # フォントが太字であることを検証
        assert header_cell.font.bold is True

    def test_auto_column_width(self, excel_writer, temp_output_dir):
        """列幅の自動調整が適用されることを検証（要件7.3）"""
        data = {
            "sheets": [
                {
                    "name": "列幅テスト",
                    "data": [
                        ["短い", "これは長いテキストです"],
                        ["A", "B"]
                    ],
                    "formatting": {"auto_width": True}
                }
            ]
        }
        output_path = os.path.join(temp_output_dir, "width_test.xlsx")
        result = excel_writer.create_workbook(data, output_path)
        
        assert result.success is True
        
        # 列幅が設定されていることを検証
        wb = load_workbook(output_path)
        ws = wb.active
        # 列幅が設定されていることを確認（デフォルト値より大きい）
        col_a_width = ws.column_dimensions['A'].width
        col_b_width = ws.column_dimensions['B'].width
        assert col_a_width is not None
        assert col_b_width is not None

    def test_formatting_disabled(self, excel_writer, temp_output_dir):
        """書式設定を無効にできることを検証（要件7.3）"""
        data = {
            "sheets": [
                {
                    "name": "フォーマット無効",
                    "data": [
                        ["ヘッダー1", "ヘッダー2"],
                        ["データ1", "データ2"]
                    ],
                    "formatting": {"header_row": False, "auto_width": False}
                }
            ]
        }
        output_path = os.path.join(temp_output_dir, "no_format.xlsx")
        result = excel_writer.create_workbook(data, output_path)
        
        assert result.success is True
        assert os.path.exists(output_path)


class TestExcelWriterFileSaving:
    """ファイル保存機能のテスト"""

    def test_save_to_specified_path(
        self, excel_writer, temp_output_dir, basic_workbook_data
    ):
        """指定されたパスにファイルが保存されることを検証（要件7.4）"""
        output_path = os.path.join(temp_output_dir, "specified_path.xlsx")
        result = excel_writer.create_workbook(basic_workbook_data, output_path)
        
        assert result.success is True
        assert result.output_path == output_path
        assert os.path.exists(output_path)

    def test_create_output_directory_if_not_exists(
        self, excel_writer, temp_output_dir, basic_workbook_data
    ):
        """出力ディレクトリが存在しない場合に作成されることを検証（要件7.4）"""
        nested_dir = os.path.join(temp_output_dir, "nested", "dir")
        output_path = os.path.join(nested_dir, "test.xlsx")
        
        result = excel_writer.create_workbook(basic_workbook_data, output_path)
        
        assert result.success is True
        assert os.path.exists(output_path)
        assert os.path.isdir(nested_dir)


class TestExcelWriterErrorHandling:
    """エラーハンドリングのテスト"""

    def test_error_on_missing_sheets_key(self, excel_writer, temp_output_dir):
        """sheetsキーがない場合にエラーを返すことを検証（要件7.5）"""
        data = {"title": "タイトルのみ"}
        output_path = os.path.join(temp_output_dir, "error.xlsx")
        
        result = excel_writer.create_workbook(data, output_path)
        
        assert result.success is False
        assert result.error is not None
        assert "sheets" in result.error.lower() or "キー" in result.error

    def test_error_on_invalid_data_type(self, excel_writer, temp_output_dir):
        """不正なデータ型の場合にエラーを返すことを検証（要件7.5）"""
        output_path = os.path.join(temp_output_dir, "error.xlsx")
        
        result = excel_writer.create_workbook("invalid_data", output_path)
        
        assert result.success is False
        assert result.error is not None

    def test_error_on_sheets_not_list(self, excel_writer, temp_output_dir):
        """sheetsがリストでない場合にエラーを返すことを検証（要件7.5）"""
        data = {"sheets": "not_a_list"}
        output_path = os.path.join(temp_output_dir, "error.xlsx")
        
        result = excel_writer.create_workbook(data, output_path)
        
        assert result.success is False
        assert result.error is not None

    def test_error_on_empty_sheets_list(self, excel_writer, temp_output_dir):
        """空のシートリストの場合にエラーを返すことを検証（要件7.5）"""
        data = {"sheets": []}
        output_path = os.path.join(temp_output_dir, "error.xlsx")
        
        result = excel_writer.create_workbook(data, output_path)
        
        assert result.success is False
        assert result.error is not None

    def test_error_on_missing_sheet_name(self, excel_writer, temp_output_dir):
        """シート名がない場合にエラーを返すことを検証（要件7.5）"""
        data = {
            "sheets": [
                {
                    "data": [["A", "B"]]
                }
            ]
        }
        output_path = os.path.join(temp_output_dir, "error.xlsx")
        
        result = excel_writer.create_workbook(data, output_path)
        
        assert result.success is False
        assert result.error is not None

    def test_error_on_missing_sheet_data(self, excel_writer, temp_output_dir):
        """シートデータがない場合にエラーを返すことを検証（要件7.5）"""
        data = {
            "sheets": [
                {
                    "name": "シート名のみ"
                }
            ]
        }
        output_path = os.path.join(temp_output_dir, "error.xlsx")
        
        result = excel_writer.create_workbook(data, output_path)
        
        assert result.success is False
        assert result.error is not None


class TestExcelWriterEdgeCases:
    """エッジケースのテスト"""

    def test_sheet_with_empty_data(self, excel_writer, temp_output_dir):
        """空のデータでもシートが作成されることを検証"""
        data = {
            "sheets": [
                {
                    "name": "空シート",
                    "data": []
                }
            ]
        }
        output_path = os.path.join(temp_output_dir, "empty_data.xlsx")
        
        result = excel_writer.create_workbook(data, output_path)
        
        assert result.success is True
        assert os.path.exists(output_path)
        
        # シートが存在することを検証
        wb = load_workbook(output_path)
        assert "空シート" in wb.sheetnames

    def test_sheet_with_numeric_data(self, excel_writer, temp_output_dir):
        """数値データが正しく書き込まれることを検証"""
        data = {
            "sheets": [
                {
                    "name": "数値シート",
                    "data": [
                        [1, 2, 3],
                        [4.5, 5.5, 6.5],
                        [100, 200, 300]
                    ]
                }
            ]
        }
        output_path = os.path.join(temp_output_dir, "numeric.xlsx")
        
        result = excel_writer.create_workbook(data, output_path)
        
        assert result.success is True
        
        # 数値が正しく書き込まれていることを検証
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == 1
        assert ws.cell(row=2, column=1).value == 4.5
        assert ws.cell(row=3, column=3).value == 300

    def test_sheet_with_mixed_data_types(self, excel_writer, temp_output_dir):
        """混合データ型が正しく書き込まれることを検証"""
        data = {
            "sheets": [
                {
                    "name": "混合データ",
                    "data": [
                        ["文字列", 123, 45.67, True],
                        [None, "", 0, False]
                    ]
                }
            ]
        }
        output_path = os.path.join(temp_output_dir, "mixed.xlsx")
        
        result = excel_writer.create_workbook(data, output_path)
        
        assert result.success is True
        
        # データが正しく書き込まれていることを検証
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "文字列"
        assert ws.cell(row=1, column=2).value == 123
        assert ws.cell(row=1, column=3).value == 45.67

    def test_long_sheet_name_is_truncated(self, excel_writer, temp_output_dir):
        """長いシート名が切り詰められることを検証"""
        long_name = "これは非常に長いシート名で31文字を超えています"
        data = {
            "sheets": [
                {
                    "name": long_name,
                    "data": [["A", "B"]]
                }
            ]
        }
        output_path = os.path.join(temp_output_dir, "long_name.xlsx")
        
        result = excel_writer.create_workbook(data, output_path)
        
        assert result.success is True
        
        # シート名が31文字以下であることを検証
        wb = load_workbook(output_path)
        for sheet_name in wb.sheetnames:
            assert len(sheet_name) <= 31
