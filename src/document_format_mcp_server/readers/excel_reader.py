"""Excel (.xlsx) file reader."""

import os
import time
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from ..utils.errors import FileNotFoundError, CorruptedFileError
from ..utils.logging_config import get_logger


# ロガーの取得
logger = get_logger(__name__)


class ExcelReader:
    """Excel (.xlsx)ファイルを読み取るクラス。"""

    def __init__(self, max_sheets: int = 100, max_file_size_mb: int = 100):
        """
        Excelリーダーを初期化する。

        Args:
            max_sheets: 処理する最大シート数（デフォルト: 100）
            max_file_size_mb: 処理する最大ファイルサイズ（MB）（デフォルト: 100）
        """
        self.max_sheets = max_sheets
        self.max_file_size_mb = max_file_size_mb

    def read_file(self, file_path: str) -> dict[str, Any]:
        """
        Excelファイルからコンテンツを抽出する。

        Args:
            file_path: 読み取るExcelファイルのパス

        Returns:
            抽出されたコンテンツを含む辞書:
            {
                "sheets": [
                    {
                        "name": str,
                        "data": [[cell_value, ...], ...],
                        "formulas": {...}
                    }
                ]
            }

        Raises:
            FileNotFoundError: ファイルが存在しない場合
            CorruptedFileError: ファイルが破損している場合
        """
        logger.info(f"Excelファイルの読み込みを開始: {file_path}")
        start_time = time.time()
        
        # ファイルの存在確認
        if not os.path.exists(file_path):
            logger.error(f"ファイルが見つかりません: {file_path}")
            raise FileNotFoundError(
                f"指定されたファイルが見つかりません: {file_path}",
                details={"file_path": file_path}
            )
        
        # ファイルサイズの検証
        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
        if file_size_mb > self.max_file_size_mb:
            logger.error(
                f"ファイルサイズが制限を超えています: {file_size_mb:.2f}MB > {self.max_file_size_mb}MB"
            )
            raise CorruptedFileError(
                f"ファイルサイズが制限を超えています: {file_size_mb:.2f}MB（最大: {self.max_file_size_mb}MB）",
                details={
                    "file_path": file_path,
                    "file_size_mb": file_size_mb,
                    "max_file_size_mb": self.max_file_size_mb
                }
            )

        try:
            # Excelファイルを開く（data_only=Falseで数式も取得）
            wb = load_workbook(file_path, data_only=False)
            
            # シート数の確認
            sheet_count = len(wb.sheetnames)
            if sheet_count > self.max_sheets:
                # 最大シート数を超える場合は警告を含めるが、最初のmax_sheetsシートを処理
                logger.warning(
                    f"シート数が制限を超えています: {sheet_count} > {self.max_sheets}。"
                    f"最初の{self.max_sheets}シートのみを処理します。"
                )
                sheets_to_process = wb.sheetnames[:self.max_sheets]
            else:
                sheets_to_process = wb.sheetnames
            
            sheets_data = []
            
            # 各シートを処理
            for sheet_name in sheets_to_process:
                sheet = wb[sheet_name]
                sheet_data = self._extract_sheet_data(sheet)
                sheets_data.append(sheet_data)
            
            result = {"sheets": sheets_data}
            
            # シート数制限の警告を追加
            if sheet_count > self.max_sheets:
                result["warning"] = f"ファイルには{sheet_count}個のシートがありますが、最初の{self.max_sheets}シートのみを処理しました。"
            
            # 処理時間を計算
            elapsed_time = time.time() - start_time
            logger.info(
                f"Excelファイルの読み込みが完了: {file_path} "
                f"(シート数: {len(sheets_data)}, 処理時間: {elapsed_time:.2f}秒)"
            )
            logger.debug(f"抽出されたデータの概要: シート数={len(sheets_data)}")
            
            return result
        
        except InvalidFileException as e:
            logger.error(
                f"Excelファイルが破損しています: {file_path}",
                exc_info=True
            )
            raise CorruptedFileError(
                f"Excelファイルが破損しているか、読み取り不可能です: {file_path}",
                details={"file_path": file_path, "error": str(e)}
            )
        except Exception as e:
            # その他の予期しないエラー
            logger.error(
                f"Excelファイルの読み取り中にエラーが発生: {file_path}",
                exc_info=True
            )
            raise CorruptedFileError(
                f"Excelファイルの読み取り中にエラーが発生しました: {file_path}",
                details={"file_path": file_path, "error": str(e)}
            )

    def _extract_sheet_data(self, sheet) -> dict[str, Any]:
        """
        シートからデータを抽出する。

        Args:
            sheet: openpyxlのWorksheetオブジェクト

        Returns:
            シートデータを含む辞書
        """
        sheet_data = {
            "name": sheet.title,
            "data": [],
            "formulas": {}
        }
        
        # シートの使用範囲を取得
        if sheet.max_row == 0 or sheet.max_column == 0:
            # 空のシート
            return sheet_data
        
        # 各行のデータを抽出
        for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
            row_data = []
            for col_idx, cell in enumerate(row, start=1):
                # セルの値を取得
                cell_value = cell.value
                
                # 数式がある場合は記録
                if cell.data_type == 'f':  # 数式セル
                    cell_address = f"{self._get_column_letter(col_idx)}{row_idx}"
                    sheet_data["formulas"][cell_address] = str(cell.value)
                    # 数式の計算結果も取得（可能な場合）
                    try:
                        # data_only=Trueで開いた場合の値を取得するため、
                        # ここでは数式文字列を保存
                        row_data.append(f"={cell.value}")
                    except Exception:
                        row_data.append(str(cell_value) if cell_value is not None else "")
                else:
                    # 通常のセル値
                    row_data.append(str(cell_value) if cell_value is not None else "")
            
            sheet_data["data"].append(row_data)
        
        return sheet_data

    def _get_column_letter(self, col_idx: int) -> str:
        """
        列番号から列文字を取得する（1 -> A, 2 -> B, ...）。

        Args:
            col_idx: 列番号（1始まり）

        Returns:
            列文字（例: "A", "B", "AA"）
        """
        result = ""
        while col_idx > 0:
            col_idx -= 1
            result = chr(col_idx % 26 + ord('A')) + result
            col_idx //= 26
        return result

