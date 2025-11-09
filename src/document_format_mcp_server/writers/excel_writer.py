"""Excel (.xlsx) file writer."""

import os
import time
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from ..utils.errors import ValidationError
from ..utils.logging_config import get_logger
from ..utils.models import WriteResult


# ロガーの取得
logger = get_logger(__name__)


class ExcelWriter:
    """Excel (.xlsx)ファイルを生成するクラス。"""

    def create_workbook(self, data: dict[str, Any], output_path: str) -> WriteResult:
        """
        構造化データからExcelファイルを生成する。

        Args:
            data: ワークブックデータ:
                {
                    "sheets": [
                        {
                            "name": str,
                            "data": [[cell_value, ...], ...],
                            "formatting": {
                                "header_row": bool (オプション),
                                "auto_width": bool (オプション)
                            } (オプション)
                        }
                    ]
                }
            output_path: 出力ファイルのパス

        Returns:
            作成されたファイルのパス

        Raises:
            ValidationError: 入力データが不正な場合
            Exception: ファイル生成中にエラーが発生した場合
        """
        logger.info(f"Excelファイルの生成を開始: {output_path}")
        start_time = time.time()

        try:
            # 入力データの検証
            self._validate_data(data)

            # 新しいワークブックを作成
            wb = Workbook()
            # デフォルトのシートを削除
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # シートを作成
            sheets_data = data.get("sheets", [])
            for sheet_data in sheets_data:
                self._create_sheet(wb, sheet_data)

            # 出力ディレクトリが存在しない場合は作成
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                try:
                    os.makedirs(output_dir, exist_ok=True)
                    logger.debug(f"出力ディレクトリを作成: {output_dir}")
                except OSError as e:
                    logger.error(f"出力ディレクトリの作成に失敗: {output_dir}", exc_info=True)
                    raise Exception(
                        f"出力ディレクトリの作成に失敗しました: {output_dir}"
                    ) from e

            # ファイルを保存
            try:
                wb.save(output_path)
            except PermissionError as e:
                logger.error(f"ファイルの保存に失敗（権限エラー）: {output_path}", exc_info=True)
                raise Exception(
                    f"ファイルの保存に失敗しました（権限エラー）: {output_path}"
                ) from e
            except OSError as e:
                logger.error(f"ファイルの保存に失敗: {output_path}", exc_info=True)
                raise Exception(
                    f"ファイルの保存に失敗しました: {output_path}"
                ) from e

            # 処理時間を計算
            elapsed_time = time.time() - start_time
            logger.info(
                f"Excelファイルの生成が完了: {output_path} "
                f"(シート数: {len(sheets_data)}, 処理時間: {elapsed_time:.2f}秒)"
            )

            # WriteResultを返す
            return WriteResult(
                success=True,
                output_path=output_path,
                url=None,
                error=None
            )

        except ValidationError as e:
            # ValidationErrorはWriteResultとして返す
            logger.error(f"Excelファイルの生成エラー（検証失敗）: {output_path}", exc_info=True)
            return WriteResult(
                success=False,
                output_path=None,
                url=None,
                error=str(e)
            )
        except Exception as e:
            logger.error(
                f"Excelファイルの生成中にエラーが発生: {output_path}",
                exc_info=True
            )
            return WriteResult(
                success=False,
                output_path=None,
                url=None,
                error=f"Excelファイルの生成中にエラーが発生しました: {str(e)}"
            )

    def _validate_data(self, data: dict[str, Any]) -> None:
        """入力データを検証する。"""
        if not isinstance(data, dict):
            raise ValidationError(
                "入力データは辞書形式である必要があります",
                details={"data_type": type(data).__name__}
            )

        if "sheets" not in data:
            raise ValidationError(
                "入力データに'sheets'キーが必要です",
                details={"keys": list(data.keys())}
            )

        sheets = data["sheets"]
        if not isinstance(sheets, list):
            raise ValidationError(
                "'sheets'はリスト形式である必要があります",
                details={"sheets_type": type(sheets).__name__}
            )

        if len(sheets) == 0:
            raise ValidationError(
                "少なくとも1つのシートが必要です",
                details={"sheet_count": 0}
            )

        # 各シートデータを検証
        for idx, sheet_data in enumerate(sheets):
            if not isinstance(sheet_data, dict):
                raise ValidationError(
                    f"シート{idx + 1}のデータは辞書形式である必要があります",
                    details={"sheet_index": idx, "data_type": type(sheet_data).__name__}
                )

            if "name" not in sheet_data:
                raise ValidationError(
                    f"シート{idx + 1}に'name'キーが必要です",
                    details={"sheet_index": idx, "keys": list(sheet_data.keys())}
                )

            if "data" not in sheet_data:
                raise ValidationError(
                    f"シート{idx + 1}に'data'キーが必要です",
                    details={"sheet_index": idx, "keys": list(sheet_data.keys())}
                )

            data_content = sheet_data["data"]
            if not isinstance(data_content, list):
                raise ValidationError(
                    f"シート{idx + 1}の'data'はリスト形式である必要があります",
                    details={"sheet_index": idx, "data_type": type(data_content).__name__}
                )

    def _create_sheet(self, wb: Workbook, sheet_data: dict[str, Any]) -> None:
        """シートを作成する。"""
        sheet_name = sheet_data.get("name", "Sheet1")
        data = sheet_data.get("data", [])
        formatting = sheet_data.get("formatting", {})

        # シートを作成
        ws = wb.create_sheet(title=sheet_name)
        logger.debug(f"シートを作成: {sheet_name}")

        if not data:
            logger.debug(f"シート '{sheet_name}' にデータがありません")
            return

        # データを書き込む
        for row_idx, row_data in enumerate(data, start=1):
            if not isinstance(row_data, list):
                row_data = [row_data]

            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = cell_value

        # フォーマットを適用
        header_row = formatting.get("header_row", True)
        auto_width = formatting.get("auto_width", True)

        # ヘッダー行のフォーマット
        if header_row and len(data) > 0:
            self._format_header_row(ws)

        # 列幅の自動調整
        if auto_width:
            self._auto_adjust_column_width(ws)

        logger.debug(
            f"シート '{sheet_name}' にデータを書き込み: "
            f"{len(data)}行 x {len(data[0]) if data else 0}列"
        )

    def _format_header_row(self, ws) -> None:
        """ヘッダー行にフォーマットを適用する。"""
        # 1行目をヘッダーとしてフォーマット
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        logger.debug("ヘッダー行のフォーマットを適用")

    def _auto_adjust_column_width(self, ws) -> None:
        """列幅を自動調整する。"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass

            # 最小幅10、最大幅50に制限
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        logger.debug("列幅の自動調整を完了")
